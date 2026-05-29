"""Streamlit web UI that reuses functions from `facturacion/gui_validation_app.py`.

The page keeps the same processing logic as the original app entrypoint and only
changes the presentation layer so the project can be shared without running a local
desktop GUI or Flask server.
"""
from __future__ import annotations

import importlib.util
import io
import os
import sys
import tempfile
import traceback
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Ensure parent workspace is on sys.path so we can import the existing module
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from facturacion import gui_validation_app as validator
from codigos import excluded_codes


DEBUG_MODE = os.environ.get("VALIDATION_DEBUG", "1") == "1"


def _debug_print(message: str):
    if DEBUG_MODE:
        print(f"[DEBUG][web_ui] {message}")


def _build_colored_table(df_display: pd.DataFrame) -> str:
    headers = list(df_display.columns)
    parts = []
    parts.append(
        '<table style="width:100%; border-collapse:collapse; font-family:Arial,sans-serif;">'
    )
    parts.append('<thead><tr>')
    for header in headers:
        parts.append(
            '<th style="border:1px solid #ddd; padding:8px; text-align:left; background:#f4f4f4;">'
            f"{header}</th>"
        )
    parts.append('</tr></thead>')
    parts.append('<tbody>')

    for _, row in df_display.iterrows():
        estado = str(row.get("Estado", "")).strip().lower()
        if estado == "ok":
            row_style = "background-color:#d4edda; color:#155724;"
        else:
            row_style = "background-color:#f8d7da; color:#721c24;"
        parts.append(f'<tr style="{row_style}">')
        for header in headers:
            value = row[header] if pd.notna(row[header]) else ""
            parts.append(
                '<td style="border:1px solid #ddd; padding:8px; vertical-align:top;">'
                f"{value}</td>"
            )
        parts.append('</tr>')

    parts.append('</tbody></table>')
    return "".join(parts)


def _excel_money(value):
    if value is None or value == "":
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    try:
        if isinstance(value, str):
            cleaned = value.replace("$", "").replace(",", "").strip()
            return float(cleaned)
        return float(value)
    except Exception:
        return value


def _expand_list_columns(df: pd.DataFrame, column_map: dict[str, str]) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    df_out = df.copy()
    for source_col, prefix in column_map.items():
        if source_col not in df_out.columns:
            continue

        max_len = 0
        normalized_values = []
        for value in df_out[source_col].tolist():
            if isinstance(value, (list, tuple, set)):
                items = list(value)
            elif value is None or (isinstance(value, float) and pd.isna(value)):
                items = []
            else:
                items = [value]
            normalized_values.append(items)
            max_len = max(max_len, len(items))

        for idx in range(max_len):
            new_col = f"{prefix} {idx + 1}"
            df_out[new_col] = [items[idx] if idx < len(items) else None for items in normalized_values]

        df_out = df_out.drop(columns=[source_col])

    return df_out


def _apply_currency_format(worksheet, header_prefixes: tuple[str, ...]) -> None:
    header_map = {}
    for cell in worksheet[1]:
        if cell.value is not None:
            header_map[str(cell.value)] = cell.column

    for header, column_index in header_map.items():
        if not header.startswith(header_prefixes):
            continue

        column_letter = get_column_letter(column_index)
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet[f"{column_letter}{row_idx}"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right")


def _apply_row_status_styles(worksheet) -> None:
    header_map = {}
    for cell in worksheet[1]:
        if cell.value is not None:
            header_map[str(cell.value)] = cell.column

    estado_col = header_map.get("Estado")
    if estado_col is None:
        return

    ok_fill = PatternFill(fill_type="solid", fgColor="D4EDDA")
    ok_font = Font(color="155724")
    error_fill = PatternFill(fill_type="solid", fgColor="F8D7DA")
    error_font = Font(color="721C24")

    for row_idx in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_idx, column=estado_col)
        estado = str(cell.value or "").strip().lower()
        if estado == "ok":
            fill = ok_fill
            font = ok_font
        else:
            fill = error_fill
            font = error_font

        for col_idx in range(1, worksheet.max_column + 1):
            current = worksheet.cell(row=row_idx, column=col_idx)
            current.fill = fill
            current.font = font


def _build_reconciliation_excel_bytes(df_transfers: pd.DataFrame | None, df_seguridad: pd.DataFrame | None, recon_mode: str) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if recon_mode == "transfers":
            sheet_name = "Transferencias"
            df_export = _expand_list_columns(
                df_transfers,
                {
                    "Neto_desprendibles": "Neto",
                    "Valores_transferencia": "Transferencia",
                },
            )
            if df_transfers is None or df_transfers.empty:
                pd.DataFrame(columns=["Identificación", "Cuenta", "Estado", "Neto 1", "Transferencia 1"]).to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                )
            else:
                df_export.to_excel(writer, sheet_name=sheet_name, index=False)
        elif recon_mode == "seguridad":
            sheet_name = "Seguridad_Social"
            df_export = df_seguridad.copy() if df_seguridad is not None else None
            if df_seguridad is None or df_seguridad.empty:
                pd.DataFrame(columns=["Identificación", "Estado", "Devengado", "IBC"]).to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                )
            else:
                if df_export is not None:
                    df_export["Devengado"] = df_export["Devengado"].apply(_excel_money)
                    df_export["IBC"] = df_export["IBC"].apply(lambda value: ", ".join(str(item) for item in value) if isinstance(value, (list, tuple, set)) else _excel_money(value))
                df_export.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            if df_transfers is not None:
                df_transfers.to_excel(writer, sheet_name="Transferencias", index=False)
            if df_seguridad is not None:
                df_seguridad.to_excel(writer, sheet_name="Seguridad_Social", index=False)

        workbook = writer.book
        if "Transferencias" in workbook.sheetnames:
            _apply_row_status_styles(workbook["Transferencias"])
            _apply_currency_format(workbook["Transferencias"], ("Neto", "Transferencia", "Devengado", "IBC"))
        if "Seguridad_Social" in workbook.sheetnames:
            _apply_row_status_styles(workbook["Seguridad_Social"])
            _apply_currency_format(workbook["Seguridad_Social"], ("Devengado", "IBC"))

    output.seek(0)
    return output.getvalue()


def _format_count(value):
    if value is None:
        return 0
    try:
        import numpy as _np
    except Exception:
        _np = None

    if _np is not None and isinstance(value, _np.generic):
        value = value.item()
    try:
        numeric = float(value)
        if numeric.is_integer():
            return int(numeric)
        return numeric
    except Exception:
        return value


def _normalize_list_like(value):
    if value is None:
        return ""
    try:
        import numpy as _np
    except Exception:
        _np = None

    if isinstance(value, (list, tuple, set)):
        parts = []
        for item in value:
            try:
                if _np is not None and isinstance(item, _np.generic):
                    item = int(item)
                parts.append(str(int(item)))
            except Exception:
                parts.append(str(item))
        return " ".join(parts)

    try:
        if _np is not None and isinstance(value, _np.generic):
            return str(int(value))
        if isinstance(value, (int, float)):
            return str(int(value))
    except Exception:
        pass

    return str(value)


def _format_dataframe(df_in: pd.DataFrame, formatter) -> pd.DataFrame:
    if df_in is None or df_in.empty:
        return df_in
    df_tmp = df_in.copy()
    list_cols = [col for col in ["Neto_desprendibles", "Valores_transferencia", "Devengado", "IBC"] if col in df_tmp.columns]

    for column in list_cols:
        def _fmt_cell(cell):
            normalized = _normalize_list_like(cell)
            try:
                return formatter(normalized)
            except Exception:
                return normalized

        df_tmp[column] = df_tmp[column].apply(_fmt_cell)
    return df_tmp


def _extract_perfiles_by_date(pdf_path: str) -> pd.DataFrame:
    registros = []
    excluded_profiles = {"none", "observaciones"}
                      


    validator_obj = validator.ServicesValidationApp.__new__(validator.ServicesValidationApp)
    with tempfile.TemporaryDirectory(prefix="web_ui_perfiles_") as _tmp_dir:
        with open(pdf_path, "rb") as pdf_handle:
            pdf_bytes = pdf_handle.read()

        tmp_pdf_path = os.path.join(_tmp_dir, "upload.pdf")
        with open(tmp_pdf_path, "wb") as temp_pdf:
            temp_pdf.write(pdf_bytes)

        import pdfplumber

        with pdfplumber.open(tmp_pdf_path) as pdf:
            for page in pdf.pages:
                for tabla in page.extract_tables() or []:
                    if not tabla or len(tabla) <= 7:
                        continue

                    header = tabla[6]
                    header_norm = [validator_obj._normalizar_busqueda(celda).replace(" ", "") if celda else "" for celda in header]
                    if "nivel/perfil" not in header_norm:
                        continue

                    idx_perfil = header_norm.index("nivel/perfil")
                    fecha_detectada = None
                    if header:
                        fecha_detectada = validator_obj._normalizar_fecha(header[-1])

                    if fecha_detectada is None:
                        continue

                    for row in tabla[7:]:
                        if len(row) <= idx_perfil:
                            continue
                        if row[4] in excluded_codes:
                            continue


                        perfil = row[idx_perfil]
                        observacion = row[-1]
                        perfil_norm = None

                        tabla_info = ""
                        try:
                            tabla_info = str(tabla[4][2])
                        except Exception:
                            tabla_info = ""

                        tabla_info_upper = tabla_info.upper()
                        if "GLOBAL" in tabla_info_upper or "NO FACTURABLE" in tabla_info_upper:
                            continue
                        


                        if isinstance(perfil, str) and observacion == "":
                            perfil = perfil.strip()
                            if perfil:
                                perfil_norm = validator_obj._normalizar_perfil(perfil)
                        elif observacion != "":

                            perfil = str(observacion).split()[-1]
                            if perfil:
                                perfil_norm = validator_obj._normalizar_perfil(perfil)

                        if not perfil_norm:
                            continue

                        cantidad = 1 / 3 if "24" in tabla_info else 1
                        registros.append(
                            {
                                "FECHA": fecha_detectada,
                                "PERFIL_NORM": perfil_norm,
                                "Nivel/Perfil": perfil_norm,
                                "PDF": cantidad,
                            }
                        )

    if not registros:
        _debug_print("No se extrajeron registros de perfiles por fecha desde el PDF.")
        return pd.DataFrame(columns=["FECHA", "PERFIL_NORM", "Nivel/Perfil", "PDF"])

    df = pd.DataFrame(registros)
    _debug_print(f"Registros de perfiles por fecha extraidos: {len(df)}")
    return df.groupby(["FECHA", "PERFIL_NORM", "Nivel/Perfil"], as_index=False)["PDF"].sum()


def _build_perfiles_table(pdf_source, excel_source) -> pd.DataFrame:
    with tempfile.TemporaryDirectory(prefix="web_ui_perfiles_table_") as tmp_dir:
        pdf_path = os.path.join(tmp_dir, "upload.pdf")
        excel_path = os.path.join(tmp_dir, "upload.xlsx")

        if hasattr(pdf_source, "getbuffer"):
            with open(pdf_path, "wb") as pdf_handle:
                pdf_handle.write(pdf_source.getbuffer())
        else:
            pdf_path = str(pdf_source)

        if hasattr(excel_source, "getbuffer"):
            with open(excel_path, "wb") as excel_handle:
                excel_handle.write(excel_source.getbuffer())
        else:
            excel_path = str(excel_source)

        df_pdf = _extract_perfiles_by_date(pdf_path)
        df_excel = _extract_excel_perfiles_by_date(excel_path)

        if df_pdf.empty:
            return pd.DataFrame(columns=["Fecha", "Nivel/Perfil", "PDF", "Excel", "Estado"])

        if df_excel.empty:
            return pd.DataFrame(columns=["Fecha", "Nivel/Perfil", "PDF", "Excel", "Estado"])

        df_merge = df_pdf.merge(
            df_excel,
            on=["FECHA", "PERFIL_NORM", "Nivel/Perfil"],
            how="outer",
        )
        df_merge["PDF"] = df_merge["PDF"].fillna(0)
        df_merge["Excel"] = df_merge["Excel"].fillna(0)

        def _estado(row):
            return "OK" if row["PDF"] == row["Excel"] else "Valores diferentes"

        df_merge["Estado"] = df_merge.apply(_estado, axis=1)
        df_merge = df_merge.rename(columns={"FECHA": "Fecha"})
        df_merge["Fecha"] = pd.to_datetime(df_merge["Fecha"], errors="coerce").dt.strftime("%Y-%m-%d")
        df_merge = df_merge[~df_merge["Nivel/Perfil"].astype(str).str.strip().str.lower().isin({"none", "incapacidad", "observaciones"})].copy()
        df_merge = df_merge[["Fecha", "Nivel/Perfil", "PDF", "Excel", "Estado"]]
        df_merge = df_merge.sort_values(["Fecha", "Nivel/Perfil"]).reset_index(drop=True)
        return df_merge


def _extract_excel_perfiles_by_date(excel_path: str) -> pd.DataFrame:
    validator_obj = validator.ServicesValidationApp.__new__(validator.ServicesValidationApp)
    excluded_profiles = {"none", "observaciones"}
    df_hist = pd.read_excel(excel_path)
    if "DESCRIPCION TARIFA" not in df_hist.columns:
        raise KeyError("El archivo Excel no contiene la columna 'DESCRIPCION TARIFA'.")

    df_niveles = df_hist[df_hist["DESCRIPCION TARIFA"].notna()].copy()
    df_niveles = df_niveles[df_niveles["DESCRIPCION TARIFA"].astype(str).str.contains("Nivel|Perfil", na=False)].copy()

    cols_fecha = [col for col in df_niveles.columns if isinstance(col, (pd.Timestamp, __import__("datetime").datetime))]
    if not cols_fecha:
        raise ValueError("No se detectaron columnas de fecha en el archivo Excel.")

    cols_id = [col for col in df_niveles.columns if col not in cols_fecha]
    df_largo = df_niveles.melt(id_vars=cols_id, value_vars=cols_fecha, var_name="FECHA", value_name="VALOR")
    df_largo["FECHA"] = pd.to_datetime(df_largo["FECHA"], errors="coerce").dt.normalize()
    df_largo = df_largo[df_largo["VALOR"].notna()].copy()
    df_largo = df_largo[df_largo["VALOR"] != 0].copy()
    df_largo["PERFIL_NORM"] = df_largo["DESCRIPCION TARIFA"].apply(validator_obj._normalizar_perfil)
    df_largo = df_largo[~df_largo["PERFIL_NORM"].astype(str).str.strip().str.lower().isin(excluded_profiles)].copy()
    df_largo["Nivel/Perfil"] = df_largo["PERFIL_NORM"]
    df_largo = df_largo.groupby(["FECHA", "PERFIL_NORM", "Nivel/Perfil"], as_index=False)["VALOR"].sum()
    df_largo = df_largo.rename(columns={"VALOR": "Excel"})
    return df_largo


def _load_payroll_module():
    payroll_path = ROOT / "mapa-de-cargos" / "gui_app.py"
    if not payroll_path.exists():
        raise FileNotFoundError("No se encontró el módulo de conciliación (mapa-de-cargos/gui_app.py).")

    spec = importlib.util.spec_from_file_location("payroll_module", str(payroll_path))
    payroll = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(payroll)
    return payroll


def _process_pagos(pdf_file, excel_file, tipo):
    _debug_print(f"Inicio _process_pagos. tipo={tipo}, pdf={getattr(pdf_file, 'name', 'N/A')}, excel={getattr(excel_file, 'name', 'N/A')}")
    validator_obj = validator.ServicesValidationApp.__new__(validator.ServicesValidationApp)

    with tempfile.TemporaryDirectory(prefix="web_ui_") as tmp_dir:
        pdf_path = os.path.join(tmp_dir, "upload.pdf")
        excel_path = os.path.join(tmp_dir, "upload.xlsx")

        with open(pdf_path, "wb") as pdf_handle:
            pdf_handle.write(pdf_file.getbuffer())
        with open(excel_path, "wb") as excel_handle:
            excel_handle.write(excel_file.getbuffer())

        rows = []

        if tipo == "perfiles":
            df_display = _build_perfiles_table(pdf_path, excel_path)
            _debug_print(f"Tabla perfiles construida. filas={len(df_display)}")
            if df_display.empty:
                return None, "No se encontraron perfiles en el PDF o no fue posible cruzarlos por fecha."
        else:
            df_pdf = validator_obj._extraer_conteo_pdf(pdf_path, tipo)
            if df_pdf is None or df_pdf.empty:
                return None, f"No se encontraron elementos válidos de tipo {tipo} en el PDF."

            pdf_agg = df_pdf.groupby(["FECHA", "TIPO DE EQUIPO"], as_index=False)["CANTIDAD"].sum()

            for fecha in sorted(pdf_agg["FECHA"].dropna().unique()):
                conteo_excel = validator_obj._extraer_conteo_excel(excel_path, fecha)
                pdf_fecha = pdf_agg[pdf_agg["FECHA"] == fecha]
                all_servicios = sorted(pdf_fecha["TIPO DE EQUIPO"].dropna().astype(str).unique().tolist())

                for servicio in all_servicios:
                    pdf_match = pdf_fecha[pdf_fecha["TIPO DE EQUIPO"] == servicio]
                    pdf_cnt = _format_count(pdf_match["CANTIDAD"].sum()) if not pdf_match.empty else 0
                    if float(pdf_cnt or 0) == 0:
                        continue
                    excel_cnt = _format_count(conteo_excel.get(servicio, 0))
                    estado = "OK" if pdf_cnt == excel_cnt else "Valores diferentes"
                    rows.append(
                        {
                            "Fecha": fecha.strftime("%Y-%m-%d") if hasattr(fecha, "strftime") else str(fecha),
                            "Servicio": servicio,
                            "PDF": pdf_cnt,
                            "Excel": excel_cnt,
                            "Estado": estado,
                        }
                    )

        if tipo != "perfiles":
            df_display = pd.DataFrame(rows)

        if df_display.empty:
            _debug_print("No hay datos para comparar en _process_pagos.")
            return None, "No se encontraron datos para comparar."

        table_html = _build_colored_table(df_display)
        _debug_print(f"Tabla final construida en _process_pagos. filas={len(df_display)}")
        return table_html, None


def _process_reconciliation(despr_files, trans_files, seguridad_files, recon_mode):
    payroll = _load_payroll_module()
    PayrollApp = payroll.PayrollReconciliationApp
    payroll_obj = PayrollApp.__new__(PayrollApp)

    with tempfile.TemporaryDirectory(prefix="web_ui_rec_") as tmp_base:
        dir_despr = os.path.join(tmp_base, "despr")
        dir_trans = os.path.join(tmp_base, "trans")
        dir_seg = os.path.join(tmp_base, "seg")
        os.makedirs(dir_despr, exist_ok=True)
        os.makedirs(dir_trans, exist_ok=True)
        os.makedirs(dir_seg, exist_ok=True)

        for uploaded_file in despr_files:
            if uploaded_file and uploaded_file.name:
                with open(os.path.join(dir_despr, uploaded_file.name), "wb") as handle:
                    handle.write(uploaded_file.getbuffer())

        if recon_mode == "transfers":
            for uploaded_file in trans_files:
                if uploaded_file and uploaded_file.name:
                    with open(os.path.join(dir_trans, uploaded_file.name), "wb") as handle:
                        handle.write(uploaded_file.getbuffer())
        else:
            for uploaded_file in seguridad_files:
                if uploaded_file and uploaded_file.name:
                    with open(os.path.join(dir_seg, uploaded_file.name), "wb") as handle:
                        handle.write(uploaded_file.getbuffer())

        df_despr = payroll_obj._process_desprendibles(dir_despr)
        df_trans = payroll_obj._process_transferencia(dir_trans) if os.listdir(dir_trans) else None
        df_seg = payroll_obj.procesar_seguridad_social(dir_seg) if os.listdir(dir_seg) else None

        reconcile_result = payroll_obj._reconcile_data(df_despr, df_trans, df_seg)
        if isinstance(reconcile_result, tuple) and len(reconcile_result) == 2:
            df_transfers, df_seguridad = reconcile_result
        else:
            df_transfers = pd.DataFrame()
            df_seguridad = pd.DataFrame()

        if (df_transfers is None or df_transfers.empty) and (df_seguridad is None or df_seguridad.empty):
            return [], "No se encontraron registros o diferencias.", df_transfers, df_seguridad

        df_display_t = df_transfers.copy() if df_transfers is not None else pd.DataFrame()
        df_display_s = df_seguridad.copy() if df_seguridad is not None else pd.DataFrame()

        df_display_t = _format_dataframe(df_display_t, payroll_obj.formatear_valores)
        df_display_s = _format_dataframe(df_display_s, payroll_obj.formatear_valores)

        parts = []
        if recon_mode == "transfers" and df_display_t is not None and not df_display_t.empty:
            parts.append(("Revisión Transferencias", _build_colored_table(df_display_t)))

        if recon_mode == "seguridad" and df_display_s is not None and not df_display_s.empty:
            parts.append(("Revisión Seguridad Social (IBC)", _build_colored_table(df_display_s)))

        if not parts:
            return [], "No se encontraron registros para el modo seleccionado.", df_transfers, df_seguridad

        return parts, None, df_transfers, df_seguridad


def main():
    st.set_page_config(page_title="Web UI", layout="wide")

    st.title("Sistema de validación y conciliación")
    st.write("Carga tus archivos para comparar PDF contra Excel o para conciliar desprendibles, transferencias y seguridad social.")

    if "perfiles_result_df" not in st.session_state:
        st.session_state.perfiles_result_df = None
    if "perfiles_result_ready" not in st.session_state:
        st.session_state.perfiles_result_ready = False
    if "recon_transfer_df" not in st.session_state:
        st.session_state.recon_transfer_df = None
    if "recon_seguridad_df" not in st.session_state:
        st.session_state.recon_seguridad_df = None
    if "recon_mode" not in st.session_state:
        st.session_state.recon_mode = None

    app_choice = st.radio(
        "Selecciona la herramienta",
        ["pagos", "mapa_transferencias", "mapa_seguridad"],
        format_func=lambda value: {
            "pagos": "Validación PDF + Excel",
            "mapa_transferencias": "Mapa de cargos - transferencias",
            "mapa_seguridad": "Mapa de cargos - seguridad social",
        }[value],
        horizontal=True,
    )

    with st.form("validation_form", clear_on_submit=False):
        if app_choice == "pagos":
            tipo = st.selectbox("Tipo de validación", ["equipos", "servicios", "perfiles"])
            pdf_file = st.file_uploader("Archivo PDF", type=["pdf"])
            excel_file = st.file_uploader("Archivo Excel", type=["xlsx", "xls"])
            submit = st.form_submit_button("Procesar")

            if submit:
                if not pdf_file:
                    st.error("Por favor sube un archivo PDF.")
                    return
                if not excel_file:
                    st.error("Por favor sube un archivo Excel.")
                    return

                try:
                    with st.spinner("Procesando archivos..."):
                        table_html, message = _process_pagos(pdf_file, excel_file, tipo)
                    if message:
                        st.info(message)
                        st.session_state.perfiles_result_df = None
                        st.session_state.perfiles_result_ready = False
                    elif table_html and tipo != "perfiles":
                        st.markdown(table_html, unsafe_allow_html=True)
                    elif tipo != "perfiles":
                        st.session_state.perfiles_result_df = None
                        st.session_state.perfiles_result_ready = False
                except Exception as exc:
                    print("[ERROR][web_ui] Procesamiento fallido en pagos")
                    traceback.print_exc()
                    st.session_state.perfiles_result_df = None
                    st.session_state.perfiles_result_ready = False
                    st.error(f"Procesamiento fallido: {exc}")

                if tipo == "perfiles" and pdf_file and excel_file:
                    try:
                        with st.spinner("Construyendo tabla por fechas..."):
                            st.session_state.perfiles_result_df = _build_perfiles_table(
                                pdf_file,
                                excel_file,
                            )
                            st.session_state.perfiles_result_ready = True
                    except Exception as exc:
                        print("[ERROR][web_ui] Construccion de tabla por fecha fallida")
                        traceback.print_exc()
                        st.session_state.perfiles_result_df = None
                        st.session_state.perfiles_result_ready = False
                        st.error(f"No se pudo construir la tabla por fecha: {exc}")

        else:
            recon_mode = "transfers" if app_choice == "mapa_transferencias" else "seguridad"
            despr_files = st.file_uploader(
                "PDFs de desprendibles",
                type=["pdf"],
                accept_multiple_files=True,
            )
            if recon_mode == "transfers":
                trans_files = st.file_uploader(
                    "PDFs de transferencias",
                    type=["pdf"],
                    accept_multiple_files=True,
                )
                seguridad_files = []
            else:
                seguridad_files = st.file_uploader(
                    "PDFs de seguridad social",
                    type=["pdf"],
                    accept_multiple_files=True,
                )
                trans_files = []

            submit = st.form_submit_button("Procesar")

            if submit:
                if not despr_files:
                    st.error("Por favor sube al menos un PDF de desprendibles.")
                    return

                if recon_mode == "transfers" and not trans_files:
                    st.error("Por favor sube al menos un PDF de transferencias para el modo 'transferencias'.")
                    return

                if recon_mode == "seguridad" and not seguridad_files:
                    st.error("Por favor sube al menos un PDF de seguridad social para el modo 'seguridad'.")
                    return

                try:
                    with st.spinner("Procesando conciliación..."):
                        parts, message, df_transfers, df_seguridad = _process_reconciliation(despr_files, trans_files, seguridad_files, recon_mode)
                        st.session_state.recon_transfer_df = df_transfers
                        st.session_state.recon_seguridad_df = df_seguridad
                        st.session_state.recon_mode = recon_mode
                    if message:
                        st.info(message)
                        st.session_state.recon_parts = []
                    else:
                        st.session_state.recon_parts = parts
                        for title, table_html in parts:
                            st.subheader(title)
                            st.markdown(table_html, unsafe_allow_html=True)
                except Exception as exc:
                    print("[ERROR][web_ui] Procesamiento de conciliacion fallido")
                    traceback.print_exc()
                    st.error(f"Procesamiento de conciliación falló: {exc}")
                    st.session_state.recon_parts = []
                    st.session_state.recon_transfer_df = None
                    st.session_state.recon_seguridad_df = None
                    st.session_state.recon_mode = None

    if st.session_state.perfiles_result_ready and isinstance(st.session_state.perfiles_result_df, pd.DataFrame):
        df_perfiles = st.session_state.perfiles_result_df
        if not df_perfiles.empty:
            st.subheader("Resultados por fecha")
            available_dates = sorted(df_perfiles["Fecha"].dropna().astype(str).unique().tolist())
            selected_date = st.selectbox(
                "Filtrar por fecha",
                ["Todas las fechas"] + available_dates,
                key="perfiles_date_filter",
            )

            if selected_date != "Todas las fechas":
                df_perfiles = df_perfiles[df_perfiles["Fecha"].astype(str) == selected_date].copy()

            st.markdown(_build_colored_table(df_perfiles), unsafe_allow_html=True)

    if "recon_parts" in st.session_state and st.session_state.recon_mode:
        df_transfers = st.session_state.recon_transfer_df
        df_seguridad = st.session_state.recon_seguridad_df

        if st.session_state.recon_parts:
            for title, table_html in st.session_state.recon_parts:
                st.subheader(title)
                st.markdown(table_html, unsafe_allow_html=True)

            excel_bytes = _build_reconciliation_excel_bytes(df_transfers, df_seguridad, st.session_state.recon_mode)
            file_name = (
                "reconciliacion_transferencias.xlsx"
                if st.session_state.recon_mode == "transfers"
                else "reconciliacion_seguridad_social.xlsx"
            )
            st.download_button(
                label="Descargar Excel con resultados",
                data=excel_bytes,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_recon_{st.session_state.recon_mode}",
            )


if __name__ == "__main__":
    main()
