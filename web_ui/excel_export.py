"""Construye un Excel (.xlsx) con los resultados para guardar/descargar.

Espeja la lógica de ``pdf_export``: una hoja por sección, encabezado con estilo
y celdas inconsistentes resaltadas en rojo. Las celdas tipo lista (mano de obra)
se aplanan: un elemento -> el valor; dos -> ``Informe: a | Lista ODS: b`` (celda
resaltada). Mantiene todo el estilo aquí para que ``app.py`` solo escriba bytes.
"""
from __future__ import annotations

import io
import re
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_DIFF_FILL = PatternFill("solid", fgColor="F8D7DA")
_DIFF_FONT = Font(color="721C24", bold=True)
_OK_FILL = PatternFill("solid", fgColor="D4EDDA")
_OK_FONT = Font(color="155724")
_WRAP_TOP = Alignment(vertical="top", wrap_text=True)


def _plano(value) -> str:
    """Texto de una celda escalar; ``""`` para nulos."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value)


def _preparar_seccion(df, source_labels):
    """Prepara ``(headers, filas)`` para una sección, expandiendo mano de obra.

    Las columnas cuyas celdas son listas (comparación de mano de obra) se separan
    en **dos columnas** —``<col> (Informe)`` y ``<col> (Lista ODS)``— para poder
    filtrar cada lado por separado en Excel, en vez de un único texto combinado.
    El resto de columnas (y las demás validaciones) quedan igual.

    Cada celda de ``filas`` es ``(texto, es_diferencia)``; ``es_diferencia`` marca
    en rojo ambas columnas del par cuando los valores no coinciden.
    """
    etiquetas = source_labels or ("Informe", "Lista ODS")
    columnas = list(df.columns)
    # Una columna se "desdobla" si alguna de sus celdas es lista (mano de obra).
    es_lista = {
        col: df[col].map(lambda v: isinstance(v, (list, tuple))).any() for col in columnas
    }

    headers = []
    for col in columnas:
        if es_lista[col]:
            headers.append(f"{col} ({etiquetas[0]})")
            headers.append(f"{col} ({etiquetas[1]})")
        else:
            headers.append(str(col))

    filas = []
    for _, row in df.iterrows():
        celdas = []
        for col in columnas:
            valor = row[col]
            if not es_lista[col]:
                celdas.append((_plano(valor), False))
                continue
            if isinstance(valor, (list, tuple)):
                inf = _plano(valor[0]) if len(valor) >= 1 else ""
                if len(valor) > 1:            # difieren: se guardó (inf, ods)
                    ods, es_diff = _plano(valor[1]), True
                else:                          # coinciden: mismo valor en ambos lados
                    ods, es_diff = inf, False
            else:                              # escalar en una columna de listas (raro)
                inf = ods = _plano(valor)
                es_diff = False
            celdas.append((inf, es_diff))
            celdas.append((ods, es_diff))
        filas.append(celdas)

    return headers, filas


def _nombre_hoja(subtitulo, usados):
    """Nombre de hoja válido (<=31 chars, sin caracteres prohibidos, único)."""
    base = re.sub(r"[\\/*?:\[\]]", " ", str(subtitulo or "Resultados")).strip() or "Resultados"
    base = base[:31]
    nombre = base
    i = 2
    while nombre in usados:
        sufijo = f" ({i})"
        nombre = base[: 31 - len(sufijo)] + sufijo
        i += 1
    usados.add(nombre)
    return nombre


def build_results_excel(titulo, archivos, secciones, source_labels=None):
    """Renderiza el Excel de resultados y devuelve sus bytes.

    Mismos parámetros que :func:`pdf_export.build_results_pdf`:
    - ``titulo``: encabezado.
    - ``archivos``: dict ``{"PDF": "...", ...}`` (puede traer listas de nombres).
    - ``secciones``: lista de ``(subtitulo, DataFrame)``; una hoja por sección.
    - ``source_labels``: par para las celdas con dos valores (mano de obra).
    """
    wb = Workbook()
    wb.remove(wb.active)
    usados = set()

    # Hoja de información (título, fecha sin hora y archivos ingresados).
    info = wb.create_sheet(_nombre_hoja("Información", usados))
    info["A1"] = titulo
    info["A1"].font = Font(bold=True, size=13)
    info["A2"] = "Generado: " + datetime.now().strftime("%Y-%m-%d")
    fila_info = 4
    if archivos:
        info[f"A{fila_info}"] = "Archivos ingresados:"
        info[f"A{fila_info}"].font = Font(bold=True)
        fila_info += 1
        for etiqueta, valor in archivos.items():
            if isinstance(valor, (list, tuple, set)):
                nombres = ", ".join(str(v) for v in valor) if valor else "—"
            else:
                nombres = str(valor) if valor else "—"
            info[f"A{fila_info}"] = f"• {etiqueta}: {nombres}"
            fila_info += 1
    info.column_dimensions["A"].width = 80

    secciones_validas = [
        (sub, df) for sub, df in secciones if isinstance(df, pd.DataFrame) and not df.empty
    ]
    if not secciones_validas:
        ws = wb.create_sheet(_nombre_hoja("Resultados", usados))
        ws["A1"] = "Sin resultados para mostrar."

    for subtitulo, df in secciones_validas:
        ws = wb.create_sheet(_nombre_hoja(subtitulo or "Resultados", usados))
        headers, filas = _preparar_seccion(df, source_labels)

        # Encabezado de la tabla.
        for col, header in enumerate(headers, start=1):
            celda = ws.cell(row=1, column=col, value=str(header))
            celda.fill = _HEADER_FILL
            celda.font = _HEADER_FONT
            celda.alignment = _WRAP_TOP

        # Una columna cuyo nombre empieza por "estado" colorea OK/diferencia.
        estado_idx = next(
            (i for i, h in enumerate(headers) if str(h).strip().lower().startswith("estado")),
            None,
        )

        for fila, celdas in enumerate(filas, start=2):
            for col, (texto, es_diff) in enumerate(celdas, start=1):
                celda = ws.cell(row=fila, column=col, value=texto)
                celda.alignment = _WRAP_TOP
                if es_diff:
                    celda.fill = _DIFF_FILL
                    celda.font = _DIFF_FONT
            if estado_idx is not None:
                est = ws.cell(row=fila, column=estado_idx + 1)
                es_ok = str(est.value).strip().lower() == "ok"
                est.fill = _OK_FILL if es_ok else _DIFF_FILL
                est.font = _OK_FONT if es_ok else _DIFF_FONT

        # Ancho de columnas aproximado al contenido del encabezado.
        for col, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col)].width = min(45, max(12, len(str(header)) + 4))
        ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
